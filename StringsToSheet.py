import os
import os.path
import re
import xml.dom.minidom
import openpyxl
import pycountry
from openpyxl.styles.fonts import Font
from openpyxl.styles.alignment import Alignment
from openpyxl.utils import get_column_letter

RES_DIR = "/Users/tonyseben/Projects/AndroidProjects/cashapp-android/cashapp/src/main/res"
COL_SLNO = 1
COL_KEY = 2
COL_DEFAULT = 3
COLUMN = 4

ROW_HEADING = 1
ROW_LANG_CODE = 2
# Empty row in row 3
ROW = 4

KEY_LIST = []

BOOK = None
SHEET = None


def getLanguages():
    print("Get languages ...")
    global RES_DIR

    languageList = []
    dirList = next(os.walk(RES_DIR))[1]

    for d in dirList:
        if d.startswith("values-"):
            lang = re.sub("values-", '', d)
            if not lang == "":
                languageList.append(lang)

    languageList.sort()
    print("Got languages : %d" % len(languageList))
    return languageList


def getStringPath(langCode):
    print("Get string path for " + langCode)
    global RES_DIR

    if not RES_DIR.endswith("/"):
        RES_DIR += "/"

    if len(langCode) > 0:
        langCode = "-" + langCode

    return RES_DIR + "values" + langCode + "/strings.xml"


def createWorkbook():
    print("Create workbook ...")
    global BOOK
    global SHEET

    BOOK = openpyxl.Workbook()
    SHEET = BOOK.active
    SHEET.row_dimensions[ROW_LANG_CODE].hidden= True


def write(row, col, value):
    global SHEET
    SHEET.cell(row, col).value = value
    SHEET.cell(row, col).alignment = Alignment(wrap_text=True)


def getRowForKey(key):
    global ROW
    global SHEET
    row = ROW

    while (not SHEET.cell(row, COL_KEY).value == ""):
        if SHEET.cell(row, COL_KEY).value == key:
            return row
        row += 1


def saveWorkbook():
    global BOOK
    BOOK.save("Strings.xls")
    print("Workbook saved!")


def isValidFile(path):
    isFile = os.path.isfile(path)
    if not isFile:
        print("Error! Path invalid : " + defaultStrPath)
    return isFile


def setupDefaultStrings():
    print("Setup sheet for default strings ...")
    global SHEET

    SHEET.cell(ROW_HEADING, COL_SLNO).value = "Sl.No"
    SHEET.cell(ROW_HEADING, COL_KEY).value = "String Key"
    SHEET.cell(ROW_HEADING, COL_DEFAULT).value = "Default"

    SHEET.cell(ROW_HEADING, COL_SLNO).font = Font(bold=True, size = "13")
    SHEET.cell(ROW_HEADING, COL_KEY).font = Font(bold=True, size = "13")
    SHEET.cell(ROW_HEADING, COL_DEFAULT).font = Font(bold=True, size = "13")

    SHEET.column_dimensions["A"].width = 6
    SHEET.column_dimensions["B"].width = 25
    SHEET.column_dimensions["C"].width = 75


def processDefaultStrings(defaultStrPath):
    print("Process default strings ...")
    global ROW
    global KEY_LIST
    row = ROW

    setupDefaultStrings()
    doc = xml.dom.minidom.parse(defaultStrPath)

    if doc.firstChild.tagName == "resources":
        stringItems = doc.getElementsByTagName("string")
        print("String count: %d" % stringItems.length)

        i = 0
        for item in stringItems:
            key = item.getAttribute("name")
            value = item.firstChild.nodeValue
            if item.getAttribute("translatable") == "false":
                value = "******"
            i += 1

            write(row, COL_SLNO, i)
            write(row, COL_KEY, key)
            write(row, COL_DEFAULT, value)

            KEY_LIST.append(key)
            row += 1

    print("Process default strings, complete.")


def processStrings(langCode):
    print("Process strings for " + langCode)
    global COLUMN
    global KEY_LIST

    path = getStringPath(langCode)
    if not isValidFile(path):
        return
    
    langCountry = re.split('-r', langCode)
    language = pycountry.languages.get(alpha_2=langCountry[0]).name
    country = ""
    if len(langCountry) > 1:
        country = " (" + pycountry.countries.get(alpha_2=langCountry[1]).name + ")"

    write(ROW_HEADING, COLUMN, language + country)
    write(ROW_LANG_CODE, COLUMN, langCode)

    SHEET.cell(ROW_HEADING, COLUMN).font = Font(bold=True, size = "13")

    doc = xml.dom.minidom.parse(path)
    if doc.firstChild.tagName == "resources":
        stringItems = doc.getElementsByTagName("string")

        for key in KEY_LIST:
            row = getRowForKey(key)

            for item in stringItems:
                if item.hasAttribute('name') and item.getAttribute('name') == key:
                    value = item.firstChild.nodeValue
                    write(row, COLUMN, value)
                    break

    colLetter = get_column_letter(COLUMN)
    SHEET.column_dimensions[colLetter].width = 75
    COLUMN += 1


# --------------------------------------#

print("################# START #################")

languagesList = getLanguages()
defaultStrPath = getStringPath("")

if isValidFile(defaultStrPath):

    createWorkbook()
    processDefaultStrings(defaultStrPath)

    for langCode in languagesList:
        processStrings(langCode)

    saveWorkbook()

print("################# FINISH #################")

