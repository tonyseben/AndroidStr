import os
import os.path
import sys
import math
import time
from openpyxl import load_workbook


BOOK_FILE = "/Users/tonyseben/Downloads/CoalitionV2Strings.xlsx"
COL_SLNO = 1
COL_KEY = 2
COL_DEFAULT = 3
COL_START = 3
COL_MAX = -1

ROW_HEADING = 1
ROW_LANG_CODE = 2
# Empty row in row 3
ROW_START = 4
ROW_MAX = -1

ROW_EMPTY_LIMIT = 20
COL_EMPTY_LIMIT = 10

BOOK = None
SHEET = None

def isValidFile(path):
    print("Validating file ...")
    isFile = os.path.isfile(path)
    if not isFile:
        print("Error! Path invalid : " + path)
    return isFile

def verifySheetFormat(path):
    print("Verifiying sheet format ...")
    global BOOK
    global SHEET
    global ROW_MAX
    global COL_MAX
    sheetName = "Translations"

    book = load_workbook(path, read_only=True)
    sheet = book[sheetName]
    if sheet is None:
        print("Error! Sheet does not exist : " + sheetName)
        return False
    else:
        BOOK = book
        SHEET = sheet
        ROW_MAX = sheet.max_row
        COL_MAX = sheet.max_column
        return True


def generateString(langColumn):
    global SHEET
    global ROW_START
    global ROW_MAX
    global COL_KEY

    lang = SHEET.cell(ROW_HEADING, langColumn).value
    langCode = SHEET.cell(ROW_LANG_CODE, langColumn).value
    langDisplay = "%s (%s)" %(lang, langCode)
    #print("Generating strings for language %s(%s) ..." % (lang, langCode))
    
    # Create directory for language.
    dirPath = "./res/values"
    if(langCode != "default"):
        dirPath += "-" + langCode
    os.makedirs(dirPath, exist_ok=True)

    strFile = open(dirPath + "/strings.xml", 'w')
    strFile.write("<?xml version=\"1.0\" encoding=\"utf-8\"?>\n<resources>\n")

    strCount = 0
    skipCount = 0
    modCount = 0
    rowEmptyCount = 0

    for row in range(ROW_START, ROW_MAX+1):
        if rowEmptyCount >= ROW_EMPTY_LIMIT:
            sys.stdout.write("\t| Aborted row scan after %d empty rows." %(ROW_EMPTY_LIMIT))
            break

        key = SHEET.cell(row, COL_KEY).value
        rawValue = SHEET.cell(row, langColumn).value

        if(key is not None and rawValue is not None):
            rowEmptyCount = 0
            value = formatString(rawValue)
            if(value != rawValue):
                modCount += 1

            if(value.strip()):
                strFile.write("\n\t<string name=\"%s\">%s</string>" %(key, value))
                strCount += 1
            else: skipCount += 1
        
        else:
            rowEmptyCount += 1

        showProgress(langDisplay , row)

    sys.stdout.write("\t| Count:%d, Skip:%d | Modified:%d\n" % (strCount, skipCount, modCount))
    strFile.write("\n\n</resources>")
    strFile.close()

def formatString(strValue):
    #print("formatString " + strValue)
    value = strValue.lstrip("\n").rstrip("\n")
    value = value.replace("\\\'", "\'")
    value = value.replace("\'", "\\\'")
    value = value.replace("\\\"", "\"")
    value = value.replace("\"", "\\\"")
    return value

def showProgress(langDisplay, currentRow):
    global ROW_START
    global ROW_MAX

    total = ROW_MAX - ROW_START
    row = currentRow - ROW_START
    progress = math.ceil(row/total*100)

    sys.stdout.write('\r* {0} {1}%'.format(langDisplay, progress))
    sys.stdout.flush()

def convertMillis(milli):
    seconds=(milli/1000)%60
    minutes=(milli/(1000*60))%60
    return "%d minutes %d seconds" % (minutes, seconds)

#**************************************************************#

startTime = int(round(time.time() * 1000))

if isValidFile(BOOK_FILE) and verifySheetFormat(BOOK_FILE):
    print("Sheet to Strings: START")
    colEmptyCount = 0

    for col in range(COL_START, COL_MAX):
        if colEmptyCount >= COL_EMPTY_LIMIT:
            print("\nPossible end of columns. Detected %d empty columns." %(COL_EMPTY_LIMIT))
            break

        if SHEET.cell(ROW_LANG_CODE, col).value is not None:
            colEmptyCount = 0
            generateString(col)
        else:
            colEmptyCount += 1

    print("Sheet to Strings: COMPLETE")

finishTime = int(round(time.time() * 1000))
time = finishTime - startTime
print("Completed in %s\n" % convertMillis(time)) 


